import requests
import json
import jsonpath
import openpyxl
from datadriven import library


def test_add_multiple_students():
    api_url = "https://thetestingworldapi.com/api/studentsDetails"
    f = open('C:/Users/ainam/Desktop/python/addnewjson.txt', 'r')
    request_json = json.loads(f.read())

    # excel code
    # vk = openpyxl.load_workbook('file path')
    # sh = vk['sheet1']
    # rows = sh.max_row

    # for i in range(2,row+1):
    # cell_fist_name =sh.cell(row=i, column=1)
    # cell_mid_name =sh.cell(row=i, column=2)
    # cell_last_name =sh.cell(row=i, column=3)
    # cell_dob =sh.cell(row=i, column=4)

    #  request_json['first_name'] = cell.fist_name.value
    #  request_json['mid_name'] =cell.mid.value
    # request_json['last_name'] =cell.last_name.value
    # request_json['dob'] =cell.dob.value
    response = requests.post(api_url, request_json)
    print(response.text)
    print(response.status_code)
    assert response.status_code == 201
