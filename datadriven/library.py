import json
import jsonpath
import requests
import openpyxl


class Common:
    def __int__(self, filenamepath, sheetname):
        global vk
        global sh
        vk = openpyxl.load_workbook(filenamepath)
        sh = vk[sheetname]

    def fetch_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        col = sh.max_column
        return col

    def fetch_key_names(self):
        c = sh.max_column
        li = []
        for i in range(1, c + 1):
            cell = sh.cell(row=1, col=i)
            li.insert(i-1, cell.value)
            return li

    def update_request_with_data(self, rownumber, jsonrequest, keylist):
        c = sh.max_column
        for i in range(1, c+1):
            cell = sh.cell(row=rownumber, column=i)
            jsonrequest[keylist-1] = cell.value
            return jsonrequest

